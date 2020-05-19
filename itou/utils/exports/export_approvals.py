import datetime
import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from django.conf import settings


from itou.approvals.models import Approval

EXPORT_DIR = f"{settings.ROOT_DIR}/exports"

FIELDS = [
    "id_pole_emploi",
    "nom",
    "prenom",
    "date_naissance",
    "numero_pass_iae",
    "date_debut_pass_iae",
    "date_fin_pass_iae",
    "code_postal",
    "code_postal_employeur",
]
DATE_FMT = "%d-%m-%Y"
EXPORT_FORMATS = ["stream", "file"]


def _approval_line(approval):
    assert approval
    return [
        approval.user.pole_emploi_id,
        approval.user.first_name,
        approval.user.last_name,
        approval.user.birthdate.strftime(DATE_FMT),
        approval.number,
        approval.start_at.strftime(DATE_FMT),
        approval.end_at.strftime(DATE_FMT),
        approval.user.post_code,
        approval.jobapplication_set.latest("created_at").to_siae.post_code,
    ]


def export_approvals(export_format=None):
    export_format = "file" if not export_format else export_format

    assert export_format in EXPORT_FORMATS, f"Unknown export format '{export_format}'"

    wb = Workbook()
    ws = wb.active

    current_dt = datetime.datetime.now()
    ws.title = "Export PASS SIAE " + current_dt.strftime(DATE_FMT)

    data = [FIELDS]

    for approval in Approval.objects.all():
        data.append(_approval_line(approval))

    max_width = [0] * (len(FIELDS) + 1)

    for i, row in enumerate(data, 1):
        for j, cell_value in enumerate(row, 1):
            max_width[j] = max(max_width[j], len(cell_value))
            ws.cell(i, j).value = cell_value
            ws.column_dimensions[get_column_letter(j)].width = max_width[j]

    if export_format == "file":
        suffix = current_dt.strftime("%d%m%Y_%H%M%S")
        path = f"{EXPORT_DIR}/export_pass_iae_{suffix}.xslx"
        wb.save(path)
        return path
    elif export_format == "stream":
        return save_virtual_workbook(wb)
