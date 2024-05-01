import datetime

from django.conf import settings
from django.db.models import F
from openpyxl import Workbook

from itou.approvals.models import Approval
from itou.utils.command import BaseCommand


class Command(BaseCommand):
    """
    Allow to export Approvals with their Suspensions/Prolongations
    """

    def add_arguments(self, parser):
        parser.add_argument(
            "--delivered-after",
            type=datetime.date.fromisoformat,
            help="Limit to PASS delivered after this date (inclusive)",
        )
        parser.add_argument(
            "--delivered-before",
            type=datetime.date.fromisoformat,
            help="Limit to PASS delivered before this date (inclusive)",
        )
        parser.add_argument(
            "--min-duration",
            type=int,
            help="Minimum PASS duration in days",
        )
        parser.add_argument(
            "--max-duration",
            type=int,
            help="Maximum PASS duration in days",
        )

    def handle(self, *args, **options):
        approvals = (
            Approval.objects.select_related("user", "user__jobseeker_profile")
            .prefetch_related("suspension_set", "prolongation_set")
            .annotate(total_duration=F("end_at") - F("start_at"))
            .order_by("-total_duration", "-start_at")
        )

        if options["delivered_after"]:
            approvals = approvals.filter(start_at__gte=options["delivered_after"])

        if options["delivered_before"]:
            approvals = approvals.filter(start_at__lte=options["delivered_before"])

        if options["min_duration"]:
            approvals = approvals.filter(total_duration__gte=datetime.timedelta(days=options["min_duration"]))

        if options["max_duration"]:
            approvals = approvals.filter(total_duration__lte=datetime.timedelta(days=options["max_duration"]))

        wb = Workbook()
        wb.iso_dates = True
        approvals_sheet = wb.active
        approvals_sheet.title = "PASS IAE"
        approvals_sheet.append(
            ["Numéro", "Début", "Fin", "Durée (jours)", "Nom", "Prénom", "NIR", "Date de naissance"]
        )
        suspensions_sheet = wb.create_sheet("Suspensions")
        suspensions_sheet.append(
            ["PASS IAE", "Début", "Fin", "Durée (jours)", "Motif", "Explications supplémentaires"]
        )
        prolongations_sheet = wb.create_sheet("Prolongations")
        prolongations_sheet.append(
            ["PASS IAE", "Début", "Fin", "Durée (jours)", "Motif", "Explications supplémentaires"]
        )

        for approval in approvals:
            approvals_sheet.append(
                [
                    approval.number,
                    approval.start_at,
                    approval.end_at,
                    approval.duration.days,
                    approval.user.last_name,
                    approval.user.first_name,
                    approval.user.jobseeker_profile.nir,
                    approval.user.birthdate if approval.user.birthdate else "",
                ]
            )
            for suspension in approval.suspension_set.all():
                suspensions_sheet.append(
                    [
                        approval.number,
                        suspension.start_at,
                        suspension.end_at,
                        suspension.duration.days,
                        suspension.get_reason_display(),
                        suspension.reason_explanation,
                    ]
                )
            for prolongation in approval.prolongation_set.all():
                prolongations_sheet.append(
                    [
                        approval.number,
                        prolongation.start_at,
                        prolongation.end_at,
                        prolongation.duration.days,
                        prolongation.get_reason_display(),
                        prolongation.reason_explanation,
                    ]
                )

        path = f"{settings.EXPORT_DIR}/export_pass_iae_expiry_{datetime.date.today()}.xlsx"
        wb.save(path)
