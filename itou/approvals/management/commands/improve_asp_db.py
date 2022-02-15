from django.core.management.base import BaseCommand
from django.db import connection
from openpyxl import load_workbook
from psycopg2 import sql  # noqa
from tqdm import tqdm

from itou.users.models import User


def get_worksheet_height(worksheet):
    # openpyxl worksheet.max_row does not return the max index with data,
    # but some max index which is not useful to us
    # https://stackoverflow.com/a/56322613
    count = 0
    for row in worksheet:
        if not all([cell.value is None for cell in row]):
            count += 1

    return count


def all_nirs_from_worksheet(worksheet, total):
    return [worksheet[f"E{row}"].value for row in range(2, total)]


def find_our_nirs(cursor, table_name="merged_approvals_poleemploiapproval"):
    """Returns a set of all the nirs we have"""
    # First we want to extract the NIRs for the pole emploi table
    # This table separates NIR and NTT/NIA, we’ll merge them
    search_sql = f"""select nir, ntt_nia from {table_name}"""

    cursor.execute(search_sql)
    res = cursor.fetchall()

    nirs = [r[0] for r in res if r[0] is not None]
    ntt_nia = [r[1] for r in res if r[1] is not None]
    set_nirs = set(nirs)
    set_ntt_nia = set(ntt_nia)
    all_nirs_and_ntt_nia = set_ntt_nia | set_nirs

    # We also include the NIRs from the user table
    search_sql = "select nir from users_user"
    cursor.execute(search_sql)
    res = cursor.fetchall()

    user_nirs = [r[0] for r in res if r[0] is not None]
    all_nirs_and_ntt_nia |= set(user_nirs)

    return all_nirs_and_ntt_nia


def display_debug_count(worksheet, cursor):
    """
    Answers the question "how many NIRs from the excel sheet do we have in our database"
    (the answer is 62%)
    """
    total = get_worksheet_height(worksheet)
    nirs = all_nirs_from_worksheet(worksheet, total)

    print("nombre de nirs distincts dans la feuille Excel:")

    print(len(nirs))

    table = "merged_approvals_poleemploiapproval"
    print(f"Nombre de ces nirs trouvés chez nous via {table}:")
    our_nirs = find_our_nirs(cursor, table)
    print(len(set(nirs) & set(our_nirs)))


class Command(BaseCommand):
    """
    ./manage.py improve_asp_db --debug --filename "exports/20220120_sans_doublons-small.xlsx"
    Complète le fichier de l'ASP (20220120_sans_doublons)
    avec les bonnes infos qui sont désormais chez nous.
    on re-remplit les colonnes de ce fichier :
        - Y (PASS IAE)
        - Z (Date de début PASS IAE)
        - AA (Date de fin PASS IAE) , en utilisant
    On utilise la colonne E (ppn_numero_inscription) qui contient le NIR pour croiser
    """

    def add_arguments(self, parser):
        parser.add_argument(
            "--filename",
            dest="filename",
            required=True,
            action="store",
            help="Absolute path of the XLSX file to import",
        )
        parser.add_argument("--debug", dest="debug", action="store_true", help="Only print the NIR counts")

    def find_agrement_data_by_nir(self, nir):
        """
        We’ve got the NIR (or possibly NTT/NIA) of a condidate,
        we want to return the agrement number, as well as it start and end date

        In order to do this, we first try to  use the (not yet deployed, hence the raw SQL)
        merged_approvals_poleemploiapproval table.
        If we do not find the candidate there, we run through the User model directly, and go through the approvals
        """
        search_sql = """select
                left(number, 12),
                start_at,
                end_at
            from merged_approvals_poleemploiapproval
            where nir=%s or ntt_nia=%s"""
        self.cursor.execute(search_sql, [nir, nir])
        res = self.cursor.fetchone()
        if res is not None:
            number, start_at, end_at = res[0], res[1], res[2]
            return (number, start_at, end_at)
        else:
            try:
                u = User.objects.get(nir=nir)  # noqa
                approval = u.approvals_wrapper.latest_approval
                if approval is not None:
                    return approval.number, approval.start_at, approval.end_at
                else:
                    raise ValueError("Approval non trouvee")

            except User.DoesNotExist:
                raise ValueError("Nir non trouve")

    def handle(self, filename, debug=False, **options):
        self.debug = debug
        self.stdout.write("Merging approvals / PASS IAE")
        self.cursor = connection.cursor()

        out = "exports/20220120_enrichi.xlsx"

        workbook = load_workbook(filename=filename)
        worksheet = workbook.active

        if debug:
            display_debug_count(worksheet, self.cursor)
        else:
            total = get_worksheet_height(worksheet)
            progressbar = tqdm(total=total)
            for row in range(2, total):
                numero_agrement_index = f"Y{row}"
                date_debut_index = f"Z{row}"
                date_fin_index = f"AA{row}"

                try:
                    nir_index = f"E{row}"
                    nir = worksheet[nir_index].value
                    number, start_at, end_at = self.find_agrement_data_by_nir(nir)

                    worksheet[numero_agrement_index] = number
                    worksheet[date_debut_index] = start_at
                    worksheet[date_fin_index] = end_at

                except ValueError:
                    worksheet[numero_agrement_index] = ""
                    worksheet[date_debut_index] = ""
                    worksheet[date_fin_index] = ""

                progressbar.update(1)

            workbook.save(out)
            progressbar.close()
