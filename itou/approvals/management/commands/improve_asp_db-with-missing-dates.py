from django.core.management.base import BaseCommand
from django.db import connection
from openpyxl import load_workbook
from psycopg2 import sql  # noqa
from tqdm import tqdm

from itou.approvals.models import Approval


def get_worksheet_height(worksheet):
    # openpyxl worksheet.max_row does not return the max index with data,
    # but some max index which is not useful to us
    # https://stackoverflow.com/a/56322613
    count = 0
    for row in worksheet:
        if not all([cell.value is None for cell in row]):
            count += 1

    return count


class Command(BaseCommand):
    """
    ./manage.py improve_asp_db-with-missing-dates --filename "imports/Date-debut et date-fin absent_rappro.xlsx"
    Complète le fichier de l'ASP
    """

    def add_arguments(self, parser):
        parser.add_argument(
            "--filename",
            dest="filename",
            required=True,
            action="store",
            help="Absolute path of the XLSX file to import",
        )

    def find_agrement_dates_by_numero_pass(self, numero):
        search_sql = """select
                        start_at,
                        end_at
                    from merged_approvals_poleemploiapproval
                    where left(number, 12)=%s"""
        self.cursor.execute(search_sql, [numero[:12]])
        res = self.cursor.fetchone()
        if res is not None:
            start_at, end_at = res[0], res[1]
            return (start_at, end_at)
        else:
            try:
                approval = Approval.objects.get(number=numero)
                return approval.start_at, approval.end_at
            except Approval.DoesNotExist:
                raise ValueError("not found")

    def handle(self, filename, debug=False, **options):
        self.debug = debug
        self.stdout.write("Filling ASP file with Itou data")
        self.cursor = connection.cursor()

        out = "exports/Date-debut et date-fin absent_rappro_enrichi.xlsx"

        workbook = load_workbook(filename=filename)
        worksheet = workbook.active

        total = get_worksheet_height(worksheet)
        progressbar = tqdm(total=total)
        for row in range(2, total):
            numero_agrement_index = f"Y{row}"
            date_debut_index = f"Z{row}"
            date_fin_index = f"AA{row}"

            try:
                numero = worksheet[numero_agrement_index].value
                start_at, end_at = self.find_agrement_dates_by_numero_pass(numero)

                worksheet[date_debut_index] = start_at
                worksheet[date_fin_index] = end_at

            except ValueError:
                worksheet[date_debut_index] = ""
                worksheet[date_fin_index] = ""

            progressbar.update(1)

        workbook.save(out)
        progressbar.close()
