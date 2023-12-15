import datetime
import logging
import time

from django.conf import settings
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from itou.job_applications.models import JobApplication, Suspension


# XLS export of approvals
# Currently used by admin site and admin command (export_approvals)

logger = logging.getLogger(__name__)

# Field names for worksheet 1 (full export of approvals)
FIELDS_WS1 = [
    "id_pole_emploi",
    "nom",
    "prenom",
    "date_naissance",
    "numero_pass_iae",
    "date_debut_pass_iae",
    "date_fin_pass_iae",
    "date_creation_pass_iae",
    "code_postal",
    "ville",
    "code_postal_employeur",
    "numero_siret",
    "raison_sociale",
    "type_siae",
    "date_debut_embauche",
    "date_fin_embauche",
]

# Field names for worksheet 2 (approvals suspensions)
FIELDS_WS2 = [
    "numero_pass_iae",
    "date_debut_suspension",
    "date_fin_suspension",
    "raison",
    "numero_siret",
    "raison_sociale",
]

# Default cell width (no more dynamic computation)
CELL_WIDTH = 50

DATE_FMT = "%d-%m-%Y"
EXPORT_FORMATS = ["stream", "file"]


def _format_date(dt):
    return dt.strftime(DATE_FMT) if dt else ""


def _format_pass_worksheet(wb):
    """
    Export of all approvals
    """
    logger.info("Loading approvals data...")

    ws = wb.active
    current_dt = datetime.datetime.now()
    ws.title = "Export PASS IAE " + current_dt.strftime(DATE_FMT)

    # Start timer
    start_counter = time.perf_counter()

    job_applications = JobApplication.objects.exclude(approval=None).select_related(
        "job_seeker", "approval", "to_company"
    )

    # Write header row
    for idx, cell_value in enumerate(FIELDS_WS1, 1):
        ws.cell(1, idx).value = cell_value

    # Write data rows
    for row_idx, ja in enumerate(job_applications.iterator(), 2):
        row = [
            ja.job_seeker.jobseeker_profile.pole_emploi_id,
            ja.job_seeker.first_name,
            ja.job_seeker.last_name,
            _format_date(ja.job_seeker.birthdate),
            ja.approval.number,
            _format_date(ja.approval.start_at),
            _format_date(ja.approval.end_at),
            _format_date(ja.approval.created_at),
            ja.job_seeker.post_code,
            ja.job_seeker.city,
            ja.to_company.post_code,
            ja.to_company.siret,
            ja.to_company.name,
            ja.to_company.kind,
            _format_date(ja.hiring_start_at),
            _format_date(ja.hiring_end_at),
        ]
        # Instead of using a temp array to store rows,
        # writing rows one by one will avoid memory issues
        for col_idx, cell_value in enumerate(row, 1):
            ws.cell(row_idx, col_idx).value = cell_value

    # Formating columns once (not in the loop)
    for idx in range(len(FIELDS_WS1)):
        ws.column_dimensions[get_column_letter(idx + 1)].width = CELL_WIDTH

    export_count = job_applications.count()

    logger.info("Exported %s approvals in %.2f sec.", export_count, time.perf_counter() - start_counter)


def _format_suspended_pass_worksheet(wb):
    """
    Suspended approvals
    """
    logger.info("Loading suspension data...")
    ws = wb.create_sheet("Suspensions PASS IAE")

    # Start timer
    start_counter = time.perf_counter()

    suspensions = Suspension.objects.all().select_related("approval", "created_by")

    # Write header row
    for idx, cell_value in enumerate(FIELDS_WS2, 1):
        ws.cell(1, idx).value = cell_value

    for idx, s in enumerate(suspensions.iterator(), 2):
        row = [
            s.approval.number,
            _format_date(s.start_at),
            _format_date(s.end_at),
            s.get_reason_display(),
            s.siae.siret,
            s.siae.name,
        ]
        for col_idx, cell_value in enumerate(row, 1):
            ws.cell(idx, col_idx).value = cell_value

    # Formating columns once (not in the loop)
    # Was dynamic, but fixed width also does the job and
    # makes code simpler
    for idx in range(len(FIELDS_WS2)):
        ws.column_dimensions[get_column_letter(idx + 1)].width = CELL_WIDTH

    export_count = suspensions.count()
    logger.info("Exported %s suspensions in %.2f sec.", export_count, time.perf_counter() - start_counter)


def export_approvals(tmp_file=None):
    """
    Main entry point. Currently used by admin site and an admin command:
        $ itou/approvals/management/commands/export_approvals.py

    `tmp_file` can be either:
        * 'None':     management command usage => save result as a file
        * valid file: admin site usage => file will be bundled in a HTTP response object

    Returns:  a valid filename for HTTP streaming (inline attachment) or storage
    """
    wb = Workbook()
    _format_pass_worksheet(wb)
    _format_suspended_pass_worksheet(wb)

    current_dt = datetime.datetime.now()
    suffix = current_dt.strftime("%d%m%Y_%H%M%S")
    filename = f"export_pass_iae_{suffix}.xlsx"

    # No streaming: management command usage
    if not tmp_file:
        path = f"{settings.EXPORT_DIR}/{filename}"
        wb.save(path)
        return path

    # Admin usage: Save tmp file for stream purposes
    wb.save(tmp_file.name)
    return filename
